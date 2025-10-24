"""
Transform raw Existing Comps Data to formatted Existing Comps sheet.

This script:
1. Reads the raw data from 'Existing Comps Data' tab
2. Filters to relevant columns
3. Sorts by Sold Price (descending)
4. Divides into quartiles
5. Applies color formatting to each quartile
6. Adds header section with quartile statistics
7. Outputs to a new Excel file
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

def transform_comps_data(input_file, output_file, sheet_name='Existing Comps Data'):
    """Transform raw comps data into formatted output."""
    
    print("=" * 80)
    print("EXISTING COMPS TRANSFORMATION SCRIPT")
    print("=" * 80)
    
    # Step 1: Read the raw data
    print(f"\n1. Reading data from '{sheet_name}' tab...")
    df_raw = pd.read_excel(input_file, sheet_name=sheet_name)
    print(f"   Loaded {len(df_raw)} rows, {len(df_raw.columns)} columns")
    
    # Step 2: Select relevant columns
    print("\n2. Selecting relevant columns...")
    columns_to_keep = [
        'Acres',
        'City',
        'DOM',
        'Garage Capacity',
        'List Price',
        'Original List Price',
        'Price Per Square Foot',
        'Sold Concessions',
        'Sold Date',
        'Sold Price',
        'Total Bedrooms',
        'Total Bathrooms',
        'Total Square Feet',
        'Year Built',
        'Property Type'
    ]
    
    df_filtered = df_raw[columns_to_keep].copy()
    print(f"   Kept {len(columns_to_keep)} columns")
    
    # Step 3: Sort by Sold Price (descending)
    print("\n3. Sorting by Sold Price (descending)...")
    df_sorted = df_filtered.sort_values('Sold Price', ascending=False, ignore_index=True)
    print(f"   Price range: ${df_sorted['Sold Price'].min():,.0f} - ${df_sorted['Sold Price'].max():,.0f}")
    
    # Step 4: Create the output workbook
    print("\n4. Creating output workbook...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Existing Comps'
    
    # Step 5: Add header section
    print("\n5. Adding header section...")
    
    # Row 2: Title
    ws['C2'] = 'Existing Sold Comps'
    ws['C2'].font = Font(bold=True, size=14)
    
    # Row 3: Subdivision name (you may want to customize this)
    ws['C3'] = 'Pagoda Grove Circle'
    ws['C3'].font = Font(bold=True)
    
    # Quartile headers
    ws['I3'] = '1st Quartile'
    ws['J3'] = '2nd Quartile'
    ws['K3'] = '3rd Quartile'
    ws['L3'] = '4th Quartile'
    for col in ['I3', 'J3', 'K3', 'L3']:
        ws[col].font = Font(bold=True)
        ws[col].alignment = Alignment(horizontal='center')
    
    # Row 5: Count
    ws['C5'] = 'Count'
    ws['D5'] = len(df_sorted)
    
    # Row 6: Quartile Size
    ws['C6'] = 'Quartile Size'
    ws['D6'] = '=D5/4'
    
    # Row 7: Criteria
    ws['C7'] = 'Criteria'
    
    # Row 8: Sold last year
    ws['C8'] = 'Sold last year'
    
    # Row 9: Location criteria (customize as needed)
    ws['C9'] = 'South of 7800, West of 2200, N'
    
    # Row 10: Property type criteria
    ws['C10'] = 'SFH, not manufactured'
    
    # Row 11: Sort info
    ws['C11'] = 'Sorted by Sold Price'
    ws['C11'].font = Font(bold=True)
    
    # Calculate quartile ranges
    # Use ceiling division for first 3 quartiles to match original behavior
    import math
    n_rows = len(df_sorted)
    q_size = math.ceil(n_rows / 4)
    q1_end = q_size
    q2_end = 2 * q_size
    q3_end = 3 * q_size
    
    # Quartile statistics labels and formulas
    stats = [
        ('Avg Sold Price', 'L'),
        ('Avg SF', 'O'),
        ('Avg Bed', 'M'),
        ('Avg Year Built', 'P'),
        ('Avg Acres', 'C'),
        ('Avg DOM', 'E'),
        ('Avg Price/SF', 'I')
    ]
    
    for idx, (label, col) in enumerate(stats, start=4):
        ws[f'H{idx}'] = label
        ws[f'H{idx}'].font = Font(bold=True)
        ws[f'I{idx}'] = f'=AVERAGE({col}$15:{col}${14 + q1_end})'
        ws[f'J{idx}'] = f'=AVERAGE({col}${15 + q1_end}:{col}${14 + q2_end})'
        ws[f'K{idx}'] = f'=AVERAGE({col}${15 + q2_end}:{col}${14 + q3_end})'
        ws[f'L{idx}'] = f'=AVERAGE({col}${15 + q3_end}:{col}${14 + n_rows})'
    
    # Step 6: Add column headers (row 14)
    print("\n6. Adding column headers...")
    for col_idx, col_name in enumerate(columns_to_keep, start=3):  # Start at column C
        cell = ws.cell(row=14, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
    
    # Step 7: Write data starting at row 15
    print("\n7. Writing data...")
    for row_idx, (_, row_data) in enumerate(df_sorted.iterrows(), start=15):
        for col_idx, col_name in enumerate(columns_to_keep, start=3):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
    
    # Step 8: Apply quartile coloring
    print("\n8. Applying quartile colors...")
    
    # Define quartile colors (light to dark green)
    colors = {
        1: 'FFE2EFD9',  # Light green (highest prices)
        2: 'FFC5E0B3',  # Medium-light green
        3: 'FFA8D08D',  # Medium green
        4: 'FF548135'   # Dark green (lowest prices)
    }
    
    # Apply colors to each quartile
    quartile_ranges = [
        (1, 15, 14 + q1_end),
        (2, 15 + q1_end, 14 + q2_end),
        (3, 15 + q2_end, 14 + q3_end),
        (4, 15 + q3_end, 14 + n_rows)
    ]
    
    for q_num, start_row, end_row in quartile_ranges:
        fill = PatternFill(start_color=colors[q_num], end_color=colors[q_num], fill_type='solid')
        print(f"   Quartile {q_num}: Rows {start_row}-{end_row} ({end_row - start_row + 1} rows)")
        
        for row in range(start_row, end_row + 1):
            for col in range(3, 3 + len(columns_to_keep)):  # Columns C onwards
                ws.cell(row=row, column=col).fill = fill
    
    # Step 9: Adjust column widths
    print("\n9. Adjusting column widths...")
    column_widths = {
        'C': 12,  # Acres
        'D': 15,  # City
        'E': 10,  # DOM
        'F': 15,  # Garage Capacity
        'G': 12,  # List Price
        'H': 18,  # Original List Price
        'I': 18,  # Price Per Square Foot
        'J': 15,  # Sold Concessions
        'K': 12,  # Sold Date
        'L': 12,  # Sold Price
        'M': 15,  # Total Bedrooms
        'N': 15,  # Total Bathrooms
        'O': 18,  # Total Square Feet
        'P': 12,  # Year Built
        'Q': 15   # Property Type
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Step 10: Save the workbook
    print(f"\n10. Saving to '{output_file}'...")
    wb.save(output_file)
    print(f"   ✓ Successfully saved!")
    
    # Summary
    print("\n" + "=" * 80)
    print("TRANSFORMATION COMPLETE")
    print("=" * 80)
    print(f"Input:  {input_file} (sheet: '{sheet_name}')")
    print(f"Output: {output_file}")
    print(f"Records processed: {len(df_sorted)}")
    print(f"Quartile sizes: Q1={q1_end}, Q2={q2_end-q1_end}, Q3={q3_end-q2_end}, Q4={n_rows-q3_end}")
    print("=" * 80)


if __name__ == '__main__':
    # Default parameters
    input_file = 'Pagoda Grove_West Jordan_Model.xlsx'
    output_file = 'Existing Comps_Transformed.xlsx'
    
    # Allow command line arguments
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    try:
        transform_comps_data(input_file, output_file)
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

